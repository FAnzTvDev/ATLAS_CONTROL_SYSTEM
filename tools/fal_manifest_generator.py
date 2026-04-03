"""
ATLAS V22 FAL PROMPT MANIFEST GENERATOR
========================================
Generates fal_prompt_manifest.xlsx — the audit sheet for every prompt hitting FAL.

Columns (exactly as specified):
  scene_id, shot_id, chain_id, is_chain_first, prev_end_frame_ref,
  characters_present, character_reference_url(s), wardrobe_tags,
  nano_mode, nano_prompt / nano_edit_prompt, negative_prompt,
  ltx_prompt, duration_s, fps, resolution, composition_hash,
  output_paths (first_frame / video)
"""

import json
import os
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --break-system-packages -q")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# Style constants
HEADER_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1a1a2e")
DATA_FONT = Font(name="Arial", size=9)
CHAIN_FILL = PatternFill("solid", fgColor="e8f0fe")  # Light blue for chained
BROLL_FILL = PatternFill("solid", fgColor="f0f0f0")  # Light gray for b-roll
ANCHOR_FILL = PatternFill("solid", fgColor="e6ffe6")  # Light green for anchors
THIN_BORDER = Border(
    left=Side(style="thin", color="cccccc"),
    right=Side(style="thin", color="cccccc"),
    top=Side(style="thin", color="cccccc"),
    bottom=Side(style="thin", color="cccccc"),
)


def generate_fal_manifest(project: str, base_path: str = None) -> str:
    """Generate the FAL prompt manifest spreadsheet."""
    if base_path is None:
        base_path = os.environ.get(
            "ATLAS_BASE",
            "/Users/quantum/Desktop/ATLAS_CONTROL_SYSTEM"
        )

    project_path = os.path.join(base_path, "pipeline_outputs", project)
    sp_path = os.path.join(project_path, "shot_plan.json")

    with open(sp_path) as f:
        sp = json.load(f)
    # T2-OR-18: bare-list guard
    if isinstance(sp, list):
        sp = {"shots": sp}

    shots = sp.get("shots", [])
    cast_map = sp.get("cast_map", {})
    chain_groups = sp.get("_chain_groups", {})

    # Load wardrobe
    wardrobe = {}
    wardrobe_path = os.path.join(project_path, "wardrobe.json")
    if os.path.exists(wardrobe_path):
        with open(wardrobe_path) as f:
            wardrobe = json.load(f)

    wb = Workbook()

    # ═══════════════════════════════════════════════════════
    # SHEET 1: FAL PROMPT MANIFEST
    # ═══════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "FAL_MANIFEST"

    headers = [
        "scene_id", "shot_id", "chain_id", "is_chain_first",
        "nano_mode", "characters_present", "character_reference_url",
        "wardrobe_tags", "nano_prompt", "delta_prompt_nano",
        "negative_prompt", "ltx_prompt", "delta_prompt_ltx",
        "duration_s", "fps", "resolution", "composition_hash",
        "actor_intent", "coverage_role", "shot_type",
        "first_frame_path", "video_path"
    ]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Write data
    for row_idx, shot in enumerate(shots, 2):
        shot_id = shot.get("shot_id", "")
        scene_id = shot.get("scene_id", "")
        chars = shot.get("characters", [])
        chain_id = shot.get("_chain_id", "")
        is_chain_first = shot.get("_is_chain_first", False)

        # Get wardrobe tags for this shot's characters
        w_tags = []
        for c in chars:
            key = f"{c}::Scene_{scene_id}"
            if key in wardrobe:
                tag = wardrobe[key].get("wardrobe_tag", "")
                if tag:
                    w_tags.append(f"{c}: {tag}")

        # Extract negative from nano_prompt
        nano = shot.get("nano_prompt", "")
        negative = ""
        if "Negative:" in nano:
            parts = nano.split("Negative:")
            negative = parts[-1].strip() if len(parts) > 1 else ""

        # Actor intent summary
        intent = shot.get("actor_intent", {})
        intent_str = ""
        if isinstance(intent, dict):
            parts = []
            for char, data in intent.items():
                if isinstance(data, dict):
                    emotion = data.get("emotion", "")
                    stature = data.get("stature", "")
                    if emotion or stature:
                        parts.append(f"{char}: {emotion}/{stature}")
            intent_str = "; ".join(parts)

        values = [
            scene_id,
            shot_id,
            chain_id,
            "YES" if is_chain_first else ("NO" if chain_id else ""),
            shot.get("_nano_mode", "text2img"),
            ", ".join(chars) if chars else "(no characters)",
            shot.get("character_reference_url", ""),
            "; ".join(w_tags) if w_tags else "",
            nano[:500],  # Truncate for readability
            shot.get("_delta_prompt_nano", ""),
            negative[:300],
            shot.get("ltx_motion_prompt", "")[:500],
            shot.get("_delta_prompt_ltx", ""),
            shot.get("duration", shot.get("duration_seconds", "")),
            24,  # Standard FPS
            "768x512",  # Standard resolution
            shot.get("_composition_hash", ""),
            intent_str,
            shot.get("coverage_role", ""),
            shot.get("shot_type", ""),
            f"first_frames/{shot_id}.jpg",
            f"videos/{shot_id}.mp4",
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Color coding
            is_broll = bool(shot.get("_broll", False) or shot.get("_no_chain", False))  # V26 DOCTRINE: suffixes are editorial, not runtime
            if is_broll:
                cell.fill = BROLL_FILL
            elif is_chain_first:
                cell.fill = ANCHOR_FILL
            elif chain_id:
                cell.fill = CHAIN_FILL

    # Set column widths
    col_widths = {
        1: 8, 2: 12, 3: 20, 4: 12, 5: 14,
        6: 30, 7: 40, 8: 35, 9: 60, 10: 45,
        11: 40, 12: 60, 13: 45, 14: 8, 15: 5,
        16: 10, 17: 14, 18: 35, 19: 14, 20: 14,
        21: 25, 22: 25
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # ═══════════════════════════════════════════════════════
    # SHEET 2: CHAIN MAP
    # ═══════════════════════════════════════════════════════
    ws2 = wb.create_sheet("CHAIN_MAP")
    chain_headers = ["chain_id", "shot_count", "scene_id", "shot_ids"]
    for col, h in enumerate(chain_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    row = 2
    for chain_id, shot_ids in sorted(chain_groups.items()):
        scene_id = shot_ids[0].split("_")[0] if shot_ids else ""
        ws2.cell(row=row, column=1, value=chain_id).font = DATA_FONT
        ws2.cell(row=row, column=2, value=len(shot_ids)).font = DATA_FONT
        ws2.cell(row=row, column=3, value=scene_id).font = DATA_FONT
        ws2.cell(row=row, column=4, value=", ".join(shot_ids)).font = DATA_FONT
        row += 1

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 80
    ws2.freeze_panes = "A2"

    # ═══════════════════════════════════════════════════════
    # SHEET 3: PREFLIGHT SUMMARY
    # ═══════════════════════════════════════════════════════
    ws3 = wb.create_sheet("PREFLIGHT")
    preflight_headers = ["check", "status", "detail", "severity"]
    for col, h in enumerate(preflight_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    # Load run report if exists
    report_path = os.path.join(project_path, "atlas_run_report.json")
    if os.path.exists(report_path):
        with open(report_path) as f:
            report = json.load(f)
        checks = report.get("phases", {}).get("preflight", {}).get("checks", [])
        for r, c in enumerate(checks, 2):
            ws3.cell(row=r, column=1, value=c.get("name", "")).font = DATA_FONT
            status_cell = ws3.cell(row=r, column=2, value="PASS" if c.get("passed") else "FAIL")
            status_cell.font = Font(name="Arial", size=9, bold=True,
                                    color="006600" if c.get("passed") else "cc0000")
            ws3.cell(row=r, column=3, value=c.get("detail", "")).font = DATA_FONT
            ws3.cell(row=r, column=4, value=c.get("severity", "")).font = DATA_FONT

    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 8
    ws3.column_dimensions["C"].width = 50
    ws3.column_dimensions["D"].width = 10
    ws3.freeze_panes = "A2"

    # ═══════════════════════════════════════════════════════
    # SHEET 4: CAST + WARDROBE
    # ═══════════════════════════════════════════════════════
    ws4 = wb.create_sheet("CAST_WARDROBE")
    cw_headers = ["character", "ai_actor", "fit_score", "scene_id", "wardrobe_tag", "locked"]
    for col, h in enumerate(cw_headers, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    row = 2
    for char_name in sorted(cast_map.keys()):
        entry = cast_map[char_name]
        actor = entry.get("ai_actor", "")
        score = entry.get("fit_score", 0)

        # Find wardrobe entries for this character
        char_wardrobe = {k: v for k, v in wardrobe.items() if v.get("character") == char_name}
        if char_wardrobe:
            for wk, wv in sorted(char_wardrobe.items()):
                ws4.cell(row=row, column=1, value=char_name).font = DATA_FONT
                ws4.cell(row=row, column=2, value=actor).font = DATA_FONT
                ws4.cell(row=row, column=3, value=score).font = DATA_FONT
                ws4.cell(row=row, column=4, value=wv.get("scene_id", "")).font = DATA_FONT
                ws4.cell(row=row, column=5, value=wv.get("wardrobe_tag", "")).font = DATA_FONT
                ws4.cell(row=row, column=6, value="YES" if wv.get("locked") else "NO").font = DATA_FONT
                row += 1
        else:
            ws4.cell(row=row, column=1, value=char_name).font = DATA_FONT
            ws4.cell(row=row, column=2, value=actor).font = DATA_FONT
            ws4.cell(row=row, column=3, value=score).font = DATA_FONT
            row += 1

    for col, w in enumerate([30, 22, 10, 10, 40, 8], 1):
        ws4.column_dimensions[get_column_letter(col)].width = w
    ws4.freeze_panes = "A2"

    # Save
    output_path = os.path.join(project_path, "fal_prompt_manifest.xlsx")
    wb.save(output_path)

    print(f"✅ FAL Manifest: {output_path}")
    print(f"   Shots: {len(shots)} | Chains: {len(chain_groups)} | Scenes: {len(set(s.get('scene_id') for s in shots))}")

    return output_path


if __name__ == "__main__":
    project = sys.argv[1] if len(sys.argv) > 1 else "ravencroft_v22"
    base = sys.argv[2] if len(sys.argv) > 2 else None
    generate_fal_manifest(project, base)
